"""
reports_generator.py
====================
Generates Stock Report, GSTR-1, GSTR-3B in PDF and Excel formats.
"""

import io
import json
import os
from datetime import datetime

# ── ReportLab ─────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# ── openpyxl ──────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Settings helper ───────────────────────────────────────────────
SETTINGS_FILE = os.path.join(os.path.dirname(__file__), 'settings.json')

def _load_settings():
    try:
        with open(SETTINGS_FILE) as f:
            return json.load(f)
    except Exception:
        return {}

def _co(s):
    """Company info shortcut."""
    return _load_settings().get('company_info', {}).get(s, '')

# ── ReportLab colour palette ──────────────────────────────────────
C_ACCENT  = colors.HexColor('#2563EB')
C_SUCCESS = colors.HexColor('#059669')
C_WARN    = colors.HexColor('#D97706')
C_DANGER  = colors.HexColor('#DC2626')
C_LIGHT   = colors.HexColor('#F8FAFC')
C_BORDER  = colors.HexColor('#E2E8F0')
C_MUTED   = colors.HexColor('#64748B')

def _rupee(val):
    try:
        return f"Rs.{float(val):,.2f}"
    except Exception:
        return "Rs.0.00"

def _num(val, decimals=2):
    try:
        return round(float(val), decimals)
    except Exception:
        return 0.0

# ─────────────────────────────────────────────────────────────────
#  COMMON: header block
# ─────────────────────────────────────────────────────────────────
def _report_header(title, subtitle, styles):
    elems = []
    company = _co('name') or 'My Company'
    gstin   = _co('gstin') or ''
    addr    = _co('address') or ''

    elems.append(Paragraph(company, ParagraphStyle(
        'CoName', fontName='Helvetica-Bold', fontSize=16,
        textColor=C_ACCENT, spaceAfter=2
    )))
    if gstin:
        elems.append(Paragraph(f"GSTIN: {gstin}", ParagraphStyle(
            'GSTIN', fontName='Helvetica', fontSize=9,
            textColor=C_MUTED, spaceAfter=1
        )))
    if addr:
        elems.append(Paragraph(addr, ParagraphStyle(
            'Addr', fontName='Helvetica', fontSize=9,
            textColor=C_MUTED, spaceAfter=6
        )))
    elems.append(HRFlowable(width='100%', thickness=1.5, color=C_ACCENT))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph(title, ParagraphStyle(
        'RepTitle', fontName='Helvetica-Bold', fontSize=13,
        textColor=colors.HexColor('#1E293B'), spaceAfter=3
    )))
    elems.append(Paragraph(subtitle, ParagraphStyle(
        'RepSub', fontName='Helvetica', fontSize=9,
        textColor=C_MUTED, spaceAfter=8
    )))
    return elems

def _tbl_style(header_col=C_ACCENT, header_text=colors.white):
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_col),
        ('TEXTCOLOR',  (0, 0), (-1, 0), header_text),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, C_LIGHT]),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.4, C_BORDER),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ])

# ─────────────────────────────────────────────────────────────────
#  STOCK REPORT — PDF
# ─────────────────────────────────────────────────────────────────
def create_stock_report_pdf(stock_data, start_date, end_date):
    buf    = io.BytesIO()
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )
    elems = _report_header(
        'Stock Report',
        f"Period: {start_date}  to  {end_date}",
        styles
    )

    # Summary strip
    total_products = len(stock_data)
    total_val      = sum(_num(r['stock_value']) for r in stock_data)
    zero_stock     = sum(1 for r in stock_data if _num(r['closing_stock']) <= 0)

    summary = [
        ['Total Products', 'Total Stock Value', 'Zero / Low Stock Items'],
        [str(total_products), _rupee(total_val), str(zero_stock)],
    ]
    st = Table(summary, colWidths=[7*cm, 7*cm, 7*cm])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), C_LIGHT),
        ('BACKGROUND', (0,1), (-1,1), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',   (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('GRID',       (0,0), (-1,-1), 0.5, C_BORDER),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elems.append(st)
    elems.append(Spacer(1, 10))

    # Main table
    headers = ['#', 'Product Name', 'HSN', 'Unit',
               'Opening Stock', 'Purchases', 'Sales',
               'Closing Stock', 'Rate (Rs.)', 'Stock Value (Rs.)']
    rows_tbl = [headers]
    for i, r in enumerate(stock_data, 1):
        closing = _num(r['closing_stock'])
        rows_tbl.append([
            str(i),
            r['name'],
            r['hsn'] or '—',
            r['unit'],
            f"{_num(r['opening_stock'], 3):g}",
            f"{_num(r['purchases'], 3):g}",
            f"{_num(r['sales'], 3):g}",
            f"{closing:g}",
            f"Rs.{_num(r['rate']):,.2f}",
            f"Rs.{_num(r['stock_value']):,.2f}",
        ])

    # Total row
    rows_tbl.append([
        '', 'TOTAL', '', '', '', '', '',
        '',
        '',
        _rupee(total_val),
    ])

    col_w = [0.8*cm, 5*cm, 1.8*cm, 1.5*cm, 2.2*cm, 2.2*cm,
             2.2*cm, 2.2*cm, 2.3*cm, 2.8*cm]
    t = Table(rows_tbl, colWidths=col_w, repeatRows=1)
    ts = _tbl_style()
    # Right-align numeric cols
    for col in [4, 5, 6, 7, 8, 9]:
        ts.add('ALIGN', (col, 1), (col, -1), 'RIGHT')
    # Highlight zero stock rows
    for i, r in enumerate(stock_data, 1):
        if _num(r['closing_stock']) <= 0:
            ts.add('TEXTCOLOR', (0, i), (-1, i), C_DANGER)
    # Total row bold
    last = len(rows_tbl) - 1
    ts.add('FONTNAME',   (0, last), (-1, last), 'Helvetica-Bold')
    ts.add('BACKGROUND', (0, last), (-1, last), C_LIGHT)
    t.setStyle(ts)
    elems.append(t)

    doc.build(elems)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
#  STOCK REPORT — EXCEL
# ─────────────────────────────────────────────────────────────────
def create_stock_report_excel(stock_data, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Report'

    company = _co('name') or 'My Company'
    gstin   = _co('gstin') or ''

    # Styles
    h_font   = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    h_fill   = PatternFill('solid', fgColor='2563EB')
    h_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tot_font = Font(name='Arial', bold=True, size=10)
    tot_fill = PatternFill('solid', fgColor='DBEAFE')
    border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    def cell_style(ws, row, col, value, bold=False, align='left',
                   fill=None, num_format=None, font_color=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='Arial', bold=bold, size=10,
                      color=font_color or '000000')
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border = border
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
        if num_format:
            c.number_format = num_format
        return c

    # Title rows
    ws.merge_cells('A1:J1')
    t = ws['A1']
    t.value = f"{company} — Stock Report"
    t.font  = Font(name='Arial', bold=True, size=14, color='1E3A5F')
    t.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    s = ws['A2']
    s.value = f"Period: {start_date} to {end_date}  |  GSTIN: {gstin}"
    s.font  = Font(name='Arial', size=9, color='64748B')
    s.alignment = Alignment(horizontal='center')

    # Header row
    headers = ['#', 'Product Name', 'HSN Code', 'Unit',
               'Opening Stock', 'Purchases', 'Sales',
               'Closing Stock', 'Avg Cost Rate (Rs.)', 'Stock Value (Rs.)']
    r = 4
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c_idx, value=h)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
        cell.border    = border
    ws.row_dimensions[r].height = 30

    # Data rows
    data_start = r + 1
    for i, row in enumerate(stock_data, 1):
        dr   = data_start + i - 1
        bg   = 'FFFFFF' if i % 2 else 'F8FAFC'
        close = _num(row['closing_stock'])
        fcolor = 'DC2626' if close <= 0 else '000000'

        cell_style(ws, dr, 1, i, align='center', fill=bg)
        cell_style(ws, dr, 2, row['name'], fill=bg, font_color=fcolor)
        cell_style(ws, dr, 3, row['hsn'] or '—', align='center', fill=bg)
        cell_style(ws, dr, 4, row['unit'], align='center', fill=bg)
        cell_style(ws, dr, 5, _num(row['opening_stock']), align='right', fill=bg, num_format='#,##0.000')
        cell_style(ws, dr, 6, _num(row['purchases']),     align='right', fill=bg, num_format='#,##0.000')
        cell_style(ws, dr, 7, _num(row['sales']),         align='right', fill=bg, num_format='#,##0.000')
        cell_style(ws, dr, 8, close,                      align='right', fill=bg, num_format='#,##0.000', font_color=fcolor)
        cell_style(ws, dr, 9, _num(row['rate']),          align='right', fill=bg, num_format='#,##0.00')
        cell_style(ws, dr, 10, _num(row['stock_value']), align='right', fill=bg, num_format='#,##0.00')

    # Total row
    tr = data_start + len(stock_data)
    last_data = tr - 1
    ws.cell(row=tr, column=2, value='TOTAL').font = tot_font
    for col in range(1, 11):
        c = ws.cell(row=tr, column=col)
        c.fill   = tot_fill
        c.border = border
        c.font   = tot_font
    total_val_cell = ws.cell(row=tr, column=10,
                              value=f'=SUM(J{data_start}:J{last_data})')
    total_val_cell.number_format = '#,##0.00'
    total_val_cell.font          = tot_font
    total_val_cell.alignment     = Alignment(horizontal='right')
    total_val_cell.fill          = tot_fill
    total_val_cell.border        = border

    # Column widths
    widths = [5, 30, 12, 8, 14, 12, 12, 14, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
#  GSTR-1 — PDF
# ─────────────────────────────────────────────────────────────────
def create_gstr1_pdf(data):
    buf    = io.BytesIO()
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )
    elems = _report_header(
        'GSTR-1 — Outward Supplies Details',
        f"Tax Period: {data['start_date']}  to  {data['end_date']}",
        styles
    )
    t = data.get('totals', {})

    def section_head(text, col=C_ACCENT):
        return Paragraph(text, ParagraphStyle(
            'SH', fontName='Helvetica-Bold', fontSize=10,
            textColor=col, spaceBefore=10, spaceAfter=4
        ))

    # ── Summary box ──────────────────────────────────────────────
    summary_rows = [
        ['Metric', 'Amount'],
        ['Total Taxable Value', _rupee(t.get('taxable', 0))],
        ['Total CGST',          _rupee(t.get('cgst', 0))],
        ['Total SGST',          _rupee(t.get('sgst', 0))],
        ['Total IGST',          _rupee(t.get('igst', 0))],
        ['Grand Total',         _rupee(t.get('grand', 0))],
        ['B2B Invoices',        str(len(data['b2b']))],
        ['B2C Invoices',        str(len(data['b2c']))],
    ]
    st = Table(summary_rows, colWidths=[8*cm, 6*cm])
    st.setStyle(_tbl_style(C_ACCENT))
    elems += [section_head('4A — Summary of Outward Supplies'), st, Spacer(1, 8)]

    # ── B2B Section ───────────────────────────────────────────────
    elems.append(section_head('B2B — Invoices to Registered Dealers'))
    if data['b2b']:
        b2b_hdr = ['Date', 'Invoice No.', 'Buyer Name', 'Buyer GSTIN',
                   'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total']
        b2b_rows = [b2b_hdr]
        for inv in data['b2b']:
            b2b_rows.append([
                inv['invoice_date'],
                inv['invoice_no'],
                inv['buyer_name'],
                inv.get('buyer_gstin') or '—',
                _rupee(inv['taxable_value']),
                _rupee(inv['total_cgst']),
                _rupee(inv['total_sgst']),
                _rupee(inv['total_igst']),
                _rupee(inv['grand_total']),
            ])
        b2b_t = Table(b2b_rows,
                      colWidths=[2*cm,3*cm,4.5*cm,3.5*cm,2.8*cm,2.2*cm,2.2*cm,2.2*cm,2.8*cm],
                      repeatRows=1)
        ts = _tbl_style()
        for col in [4,5,6,7,8]:
            ts.add('ALIGN', (col,1), (col,-1), 'RIGHT')
        b2b_t.setStyle(ts)
        elems.append(b2b_t)
    else:
        elems.append(Paragraph('No B2B invoices in this period.', styles['Normal']))
    elems.append(Spacer(1, 8))

    # ── B2C Section ───────────────────────────────────────────────
    elems.append(section_head('B2C — Invoices to Unregistered / Consumers'))
    if data['b2c']:
        b2c_hdr = ['Date', 'Invoice No.', 'Buyer Name', 'Taxable Value',
                   'CGST', 'SGST', 'IGST', 'Total']
        b2c_rows = [b2c_hdr]
        for inv in data['b2c']:
            b2c_rows.append([
                inv['invoice_date'],
                inv['invoice_no'],
                inv['buyer_name'],
                _rupee(inv['taxable_value']),
                _rupee(inv['total_cgst']),
                _rupee(inv['total_sgst']),
                _rupee(inv['total_igst']),
                _rupee(inv['grand_total']),
            ])
        b2c_t = Table(b2c_rows,
                      colWidths=[2*cm,3*cm,5.5*cm,3*cm,2.5*cm,2.5*cm,2.5*cm,3*cm],
                      repeatRows=1)
        ts = _tbl_style(header_col=C_SUCCESS)
        for col in [3,4,5,6,7]:
            ts.add('ALIGN', (col,1), (col,-1), 'RIGHT')
        b2c_t.setStyle(ts)
        elems.append(b2c_t)
    else:
        elems.append(Paragraph('No B2C invoices in this period.', styles['Normal']))
    elems.append(Spacer(1, 8))

    # ── HSN Summary ───────────────────────────────────────────────
    elems.append(section_head('HSN Summary (Section 12)'))
    if data['hsn_summary']:
        hsn_hdr = ['HSN Code', 'GST Rate', 'Total Qty', 'Taxable Value', 'Tax Amount']
        hsn_rows = [hsn_hdr]
        for h in data['hsn_summary']:
            hsn_rows.append([
                h['hsn'] or '—',
                f"{_num(h['gst_rate'])}%",
                f"{_num(h['total_qty'], 3):g}",
                _rupee(h['taxable_value']),
                _rupee(h['total_tax']),
            ])
        hsn_t = Table(hsn_rows, colWidths=[3*cm,2.5*cm,3*cm,5*cm,4*cm], repeatRows=1)
        ts = _tbl_style(header_col=C_WARN, header_text=colors.white)
        for col in [2,3,4]:
            ts.add('ALIGN', (col,1), (col,-1), 'RIGHT')
        hsn_t.setStyle(ts)
        elems.append(hsn_t)
    else:
        elems.append(Paragraph('No HSN data available.', styles['Normal']))

    doc.build(elems)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
#  GSTR-1 — EXCEL
# ─────────────────────────────────────────────────────────────────
def create_gstr1_excel(data):
    wb = Workbook()
    company = _co('name') or 'My Company'
    gstin   = _co('gstin') or ''
    period  = f"{data['start_date']} to {data['end_date']}"

    def _header_row(ws, headers, fill_color='2563EB'):
        r = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor=fill_color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        ws.row_dimensions[r].height = 28
        return r

    def _data_row(ws, vals, row_num, num_cols=None, fill=None):
        r = ws.max_row + 1
        bg = 'FFFFFF' if row_num % 2 else 'F8FAFC'
        if fill: bg = fill
        bdr = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(name='Arial', size=9)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.border    = bdr
            if isinstance(v, (int, float)) and (num_cols is None or c in num_cols):
                cell.number_format = '#,##0.00'
                cell.alignment     = Alignment(horizontal='right')

    def _title(ws, title):
        r = ws.max_row + 1 if ws.max_row > 0 else 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(row=r, column=1, value=title)
        c.font      = Font(name='Arial', bold=True, size=11, color='1E3A5F')
        c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 20

    # ── Sheet 1: B2B ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'B2B'
    _title(ws1, f"{company} | GSTR-1 B2B | {period}")
    ws1.append([])
    _header_row(ws1, ['Date','Invoice No.','Buyer Name','Buyer GSTIN',
                       'Taxable Value','CGST','SGST','IGST','Grand Total'])
    for i, inv in enumerate(data['b2b'], 1):
        _data_row(ws1, [
            inv['invoice_date'], inv['invoice_no'],
            inv['buyer_name'], inv.get('buyer_gstin') or '',
            _num(inv['taxable_value']), _num(inv['total_cgst']),
            _num(inv['total_sgst']), _num(inv['total_igst']),
            _num(inv['grand_total']),
        ], i, num_cols={5,6,7,8,9})
    for col, w in enumerate([12,18,25,18,14,10,10,10,14], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A4'

    # ── Sheet 2: B2C ──────────────────────────────────────────────
    ws2 = wb.create_sheet('B2C')
    _title(ws2, f"{company} | GSTR-1 B2C | {period}")
    ws2.append([])
    _header_row(ws2, ['Date','Invoice No.','Buyer Name',
                       'Taxable Value','CGST','SGST','IGST','Grand Total'],
                fill_color='059669')
    for i, inv in enumerate(data['b2c'], 1):
        _data_row(ws2, [
            inv['invoice_date'], inv['invoice_no'], inv['buyer_name'],
            _num(inv['taxable_value']), _num(inv['total_cgst']),
            _num(inv['total_sgst']), _num(inv['total_igst']),
            _num(inv['grand_total']),
        ], i, num_cols={4,5,6,7,8})
    for col, w in enumerate([12,18,30,14,10,10,10,14], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A4'

    # ── Sheet 3: HSN Summary ──────────────────────────────────────
    ws3 = wb.create_sheet('HSN Summary')
    _title(ws3, f"{company} | GSTR-1 HSN Summary | {period}")
    ws3.append([])
    _header_row(ws3, ['HSN Code','GST Rate %','Total Qty',
                       'Taxable Value','Total Tax'],
                fill_color='D97706')
    for i, h in enumerate(data['hsn_summary'], 1):
        _data_row(ws3, [
            h['hsn'] or '', _num(h['gst_rate']),
            _num(h['total_qty'], 3),
            _num(h['taxable_value']), _num(h['total_tax']),
        ], i, num_cols={4,5})
    for col, w in enumerate([14,12,12,18,14], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
#  GSTR-3B — PDF
# ─────────────────────────────────────────────────────────────────
def create_gstr3b_pdf(data):
    buf    = io.BytesIO()
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )
    elems = _report_header(
        'GSTR-3B — Monthly Summary Return',
        f"Tax Period: {data['start_date']}  to  {data['end_date']}",
        styles
    )
    t = data.get('tax_totals', {})

    # ── 3.1 Outward Supplies ─────────────────────────────────────
    def sh(txt, col=C_ACCENT):
        return Paragraph(txt, ParagraphStyle(
            's', fontName='Helvetica-Bold', fontSize=10,
            textColor=col, spaceBefore=10, spaceAfter=4
        ))

    elems.append(sh('3.1 — Outward Taxable Supplies'))
    main_rows = [
        ['Description', 'Total Invoices', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total Tax'],
        [
            'Total Outward Supplies',
            str(t.get('invoice_count', 0)),
            _rupee(t.get('total_taxable', 0)),
            _rupee(t.get('total_cgst', 0)),
            _rupee(t.get('total_sgst', 0)),
            _rupee(t.get('total_igst', 0)),
            _rupee(t.get('total_gst', 0)),
        ],
    ]
    mt = Table(main_rows, colWidths=[5*cm,2.5*cm,3*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm])
    ts = _tbl_style()
    for col in [1,2,3,4,5,6]:
        ts.add('ALIGN', (col,1), (col,-1), 'RIGHT')
    ts.add('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold')
    mt.setStyle(ts)
    elems.append(mt)
    elems.append(Spacer(1, 10))

    # ── Rate-wise breakup ─────────────────────────────────────────
    elems.append(sh('Rate-wise Tax Breakup'))
    rw_hdr = ['GST Rate', 'Taxable Value', 'CGST', 'SGST', 'IGST (est.)', 'Total Tax']
    rw_rows = [rw_hdr]
    for r in data.get('rate_wise', []):
        tax_total = _num(r.get('cgst', 0)) + _num(r.get('sgst', 0))
        rw_rows.append([
            f"{_num(r['gst_rate'])}%",
            _rupee(r.get('taxable_value', 0)),
            _rupee(r.get('cgst', 0)),
            _rupee(r.get('sgst', 0)),
            _rupee(r.get('igst_total', 0)),
            _rupee(tax_total),
        ])
    rw_t = Table(rw_rows, colWidths=[2.5*cm,4*cm,3*cm,3*cm,3*cm,3*cm])
    ts = _tbl_style(header_col=C_WARN)
    for col in [1,2,3,4,5]:
        ts.add('ALIGN', (col,1), (col,-1), 'RIGHT')
    rw_t.setStyle(ts)
    elems.append(rw_t)
    elems.append(Spacer(1, 10))

    # ── 3.2 Tax liability summary ─────────────────────────────────
    elems.append(sh('3.2 — Tax Liability Summary'))
    tl_rows = [
        ['Tax Head', 'Amount'],
        ['CGST Payable',       _rupee(t.get('total_cgst', 0))],
        ['SGST Payable',       _rupee(t.get('total_sgst', 0))],
        ['IGST Payable',       _rupee(t.get('total_igst', 0))],
        ['Total Tax Payable',  _rupee(t.get('total_gst', 0))],
    ]
    tl_t = Table(tl_rows, colWidths=[10*cm, 5*cm])
    ts = _tbl_style(header_col=C_DANGER)
    ts.add('ALIGN', (1,1), (1,-1), 'RIGHT')
    ts.add('FONTNAME', (0,4), (-1,4), 'Helvetica-Bold')
    ts.add('BACKGROUND', (0,4), (-1,4), colors.HexColor('#FEF2F2'))
    tl_t.setStyle(ts)
    elems.append(tl_t)

    # ── Section 4: Eligible ITC ──────────────────────────────
    # ITC section ALWAYS shows — even if 0 (no purchases)
    # Net Tax Payable ALWAYS shows = Output Tax - ITC
    itc = data.get('itc', {})
    ei  = itc.get('eligible_itc', {}) if itc else {}
    rcm = itc.get('rcm_itc',      {}) if itc else {}

    elems.append(sh('4 — Eligible Input Tax Credit (ITC) from Purchases', col=C_SUCCESS))

    itc_rows = [
        ['ITC Type', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total ITC'],
        [
            'Eligible ITC (Regular Purchases)',
            _rupee(ei.get('taxable', 0)),
            _rupee(ei.get('cgst', 0)),
            _rupee(ei.get('sgst', 0)),
            _rupee(ei.get('igst', 0)),
            _rupee(ei.get('total_tax', 0)),
        ],
        [
            'RCM ITC (Reverse Charge)',
            '—',
            _rupee(rcm.get('cgst', 0)),
            _rupee(rcm.get('sgst', 0)),
            _rupee(rcm.get('igst', 0)),
            _rupee(rcm.get('total_tax', 0)),
        ],
    ]
    total_itc = _num(ei.get('total_tax', 0)) + _num(rcm.get('total_tax', 0))
    itc_rows.append([
        'Total ITC Available', '—', '—', '—', '—', _rupee(total_itc)
    ])
    itc_t = Table(itc_rows, colWidths=[5.5*cm,3*cm,2.5*cm,2.5*cm,2.5*cm,3*cm])
    ts = _tbl_style(header_col=C_SUCCESS)
    for col in [1,2,3,4,5]:
        ts.add('ALIGN', (col,1), (col,-1), 'RIGHT')
    ts.add('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold')
    ts.add('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#ECFDF5'))
    itc_t.setStyle(ts)
    elems.append(itc_t)
    elems.append(Spacer(1, 10))

    # ── Net Tax Payable — ALWAYS shown, ITC subtracted ────────────
    elems.append(sh('Net Tax Payable (Output − ITC)', col=C_DANGER))
    out_cgst  = _num(t.get('total_cgst', 0))
    out_sgst  = _num(t.get('total_sgst', 0))
    out_igst  = _num(t.get('total_igst', 0))
    out_total = _num(t.get('total_gst',  0))
    itc_cgst  = _num(ei.get('cgst', 0)) + _num(rcm.get('cgst', 0))
    itc_sgst  = _num(ei.get('sgst', 0)) + _num(rcm.get('sgst', 0))
    itc_igst  = _num(ei.get('igst', 0)) + _num(rcm.get('igst', 0))
    itc_total = total_itc

    # ── GST ITC Cross-Utilization (as per GST Act) ───────────────
    # Rule: IGST credit can offset IGST first, then CGST, then SGST
    #       CGST credit can offset CGST first, then IGST
    #       SGST credit can offset SGST first, then IGST

    # Step 1: Use same-head ITC first
    rem_igst_itc = itc_igst
    rem_cgst_itc = itc_cgst
    rem_sgst_itc = itc_sgst

    # IGST vs IGST
    net_igst      = max(0, out_igst - rem_igst_itc)
    rem_igst_itc  = max(0, rem_igst_itc - out_igst)

    # CGST vs CGST
    net_cgst      = max(0, out_cgst - rem_cgst_itc)
    rem_cgst_itc  = max(0, rem_cgst_itc - out_cgst)

    # SGST vs SGST
    net_sgst      = max(0, out_sgst - rem_sgst_itc)
    rem_sgst_itc  = max(0, rem_sgst_itc - out_sgst)

    # Step 2: Use remaining IGST ITC against CGST, then SGST
    if rem_igst_itc > 0 and net_cgst > 0:
        used         = min(rem_igst_itc, net_cgst)
        net_cgst    -= used
        rem_igst_itc -= used

    if rem_igst_itc > 0 and net_sgst > 0:
        used         = min(rem_igst_itc, net_sgst)
        net_sgst    -= used
        rem_igst_itc -= used

    # Step 3: Use remaining CGST ITC against IGST
    if rem_cgst_itc > 0 and net_igst > 0:
        used         = min(rem_cgst_itc, net_igst)
        net_igst    -= used

    # Step 4: Use remaining SGST ITC against IGST
    if rem_sgst_itc > 0 and net_igst > 0:
        used         = min(rem_sgst_itc, net_igst)
        net_igst    -= used

    net_cgst  = round(net_cgst,  2)
    net_sgst  = round(net_sgst,  2)
    net_igst  = round(net_igst,  2)
    net_total = round(net_cgst + net_sgst + net_igst, 2)

    net_rows = [
        ['', 'CGST', 'SGST', 'IGST', 'Total'],
        ['Output Tax (Sales)',    _rupee(out_cgst),  _rupee(out_sgst),  _rupee(out_igst),  _rupee(out_total)],
        ['Less: ITC Available',   _rupee(itc_cgst),  _rupee(itc_sgst),  _rupee(itc_igst),  _rupee(itc_total)],
        ['Net Tax Payable (Cash)', _rupee(net_cgst),  _rupee(net_sgst),  _rupee(net_igst),  _rupee(net_total)],
    ]
    net_t = Table(net_rows, colWidths=[6*cm,3*cm,3*cm,3*cm,3*cm])
    ts2 = _tbl_style(header_col=C_DANGER)
    for col in [1,2,3,4]:
        ts2.add('ALIGN', (col,1), (col,-1), 'RIGHT')
    ts2.add('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold')
    ts2.add('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#FEF2F2'))
    net_t.setStyle(ts2)
    elems.append(net_t)

    # Note
    elems.append(Spacer(1, 12))
    elems.append(Paragraph(
        '* IGST breakup shown is approximate. Cross-check with actual interstate invoices. '
        'ITC eligible only on purchases with valid tax invoices from GST-registered suppliers. '
        'File GSTR-3B on the GST portal using these figures.',
        ParagraphStyle('note', fontName='Helvetica-Oblique', fontSize=8,
                       textColor=C_MUTED)
    ))

    doc.build(elems)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
#  GSTR-3B — EXCEL
# ─────────────────────────────────────────────────────────────────
def create_gstr3b_excel(data):
    wb      = Workbook()
    ws      = wb.active
    ws.title = 'GSTR-3B'
    company  = _co('name') or 'My Company'
    period   = f"{data['start_date']} to {data['end_date']}"
    t        = data.get('tax_totals', {})

    bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _w(row, col, val, bold=False, fill=None, align='left', fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Arial', bold=bold, size=10)
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border    = bdr
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
        if fmt:
            c.number_format = fmt
        return c

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'].value     = f"{company} — GSTR-3B Summary | {period}"
    ws['A1'].font      = Font(name='Arial', bold=True, size=14, color='1E3A5F')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Section 3.1
    r = 3
    ws.merge_cells(f'A{r}:G{r}')
    ws[f'A{r}'].value = '3.1 — Outward Taxable Supplies'
    ws[f'A{r}'].font  = Font(name='Arial', bold=True, size=11, color='2563EB')

    r = 4
    for ci, h in enumerate(['Description','Invoices','Taxable Value','CGST','SGST','IGST','Total Tax'], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill  = PatternFill('solid', fgColor='2563EB')
        c.alignment = Alignment(horizontal='center')
        c.border = bdr

    r = 5
    vals = ['Total Outward Supplies', t.get('invoice_count',0),
            _num(t.get('total_taxable',0)), _num(t.get('total_cgst',0)),
            _num(t.get('total_sgst',0)),   _num(t.get('total_igst',0)),
            _num(t.get('total_gst',0))]
    for ci, v in enumerate(vals, 1):
        _w(r, ci, v, bold=True, fill='DBEAFE',
           align='right' if ci > 1 else 'left',
           fmt='#,##0.00' if ci > 2 else None)

    # Section: Rate-wise
    r = 7
    ws.merge_cells(f'A{r}:G{r}')
    ws[f'A{r}'].value = 'Rate-wise Breakup'
    ws[f'A{r}'].font  = Font(name='Arial', bold=True, size=11, color='D97706')

    r = 8
    for ci, h in enumerate(['GST Rate','Taxable Value','CGST','SGST','IGST (est.)','Total Tax',''], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill  = PatternFill('solid', fgColor='D97706')
        c.alignment = Alignment(horizontal='center')
        c.border = bdr

    r = 9
    for row in data.get('rate_wise', []):
        tax = _num(row.get('cgst',0)) + _num(row.get('sgst',0))
        for ci, (v, fmt) in enumerate([
            (f"{_num(row['gst_rate'])}%", None),
            (_num(row.get('taxable_value',0)), '#,##0.00'),
            (_num(row.get('cgst',0)),          '#,##0.00'),
            (_num(row.get('sgst',0)),           '#,##0.00'),
            (_num(row.get('igst_total',0)),     '#,##0.00'),
            (tax,                               '#,##0.00'),
        ], 1):
            _w(r, ci, v, align='right' if ci > 1 else 'center', fmt=fmt)
        r += 1

    # Section 3.2
    r += 1
    ws.merge_cells(f'A{r}:G{r}')
    ws[f'A{r}'].value = '3.2 — Tax Liability Summary'
    ws[f'A{r}'].font  = Font(name='Arial', bold=True, size=11, color='DC2626')

    r += 1
    for ci, h in enumerate(['Tax Head', 'Amount Payable'], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill  = PatternFill('solid', fgColor='DC2626')
        c.alignment = Alignment(horizontal='center')
        c.border = bdr

    r += 1
    for label, key in [('CGST','total_cgst'),('SGST','total_sgst'),
                        ('IGST','total_igst'),('Total Tax','total_gst')]:
        bold = label == 'Total Tax'
        fill = 'FEF2F2' if bold else 'FFFFFF'
        _w(r, 1, label, bold=bold, fill=fill)
        _w(r, 2, _num(t.get(key, 0)), bold=bold, fill=fill,
           align='right', fmt='#,##0.00')
        r += 1

    # Column widths
    for col, w in enumerate([25,14,16,12,12,14,12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  GSTR-2B / PURCHASE REGISTER REPORT
# ─────────────────────────────────────────────────────────────────────────────

def create_gstr2b_pdf(data):
    """
    GSTR-2B style Purchase Register PDF.
    Shows all purchase bills with Supplier, GSTIN, Bill No, Date,
    Place of Supply, RCM flag, Taxable, CGST, SGST, IGST, Total Tax.
    """
    from reportlab.lib import colors as rl_colors
    buf    = io.BytesIO()
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm
    )

    elems = _report_header(
        'Purchase Register — ITC Statement',
        f"Period: {data['start_date']}  to  {data['end_date']}",
        styles
    )

    purchases = data.get('purchases', [])
    totals    = data.get('totals', {})

    # ── Summary box ──────────────────────────────────────────────
    co    = _co
    s_hdr = ParagraphStyle('sh', fontName='Helvetica-Bold', fontSize=9,
                           textColor=C_ACCENT, spaceAfter=4)
    elems.append(Paragraph('Summary', s_hdr))

    sum_rows = [
        ['Total Bills', 'Total Taxable', 'Total CGST', 'Total SGST', 'Total IGST', 'Total Tax', 'Grand Total'],
        [
            str(totals.get('bill_count', 0)),
            _rupee(totals.get('taxable',   0)),
            _rupee(totals.get('cgst',      0)),
            _rupee(totals.get('sgst',      0)),
            _rupee(totals.get('igst',      0)),
            _rupee(totals.get('total_tax', 0)),
            _rupee(totals.get('grand',     0)),
        ]
    ]
    s_tbl = Table(sum_rows, colWidths=[2.5*cm,3.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm,3*cm])
    ts = _tbl_style()
    for col in range(1, 7):
        ts.add('ALIGN', (col, 1), (col, -1), 'RIGHT')
    ts.add('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold')
    s_tbl.setStyle(ts)
    elems.append(s_tbl)
    elems.append(Spacer(1, 10))

    # ── Purchase Detail Table ─────────────────────────────────────
    elems.append(Paragraph('Purchase Detail', s_hdr))

    hdr = ['#', 'Supplier', 'GSTIN', 'Bill No.', 'Date',
           'Place of Supply', 'RCM', 'Taxable', 'CGST', 'SGST', 'IGST', 'Total Tax']
    rows = [hdr]

    small = ParagraphStyle('sm', fontName='Helvetica', fontSize=7, leading=9)
    bold7 = ParagraphStyle('b7', fontName='Helvetica-Bold', fontSize=7, leading=9)

    for i, p in enumerate(purchases, 1):
        rcm = 'Yes' if p.get('reverse_charge') else 'No'
        is_igst = float(p.get('igst_amount', 0) or 0) > 0
        rows.append([
            str(i),
            Paragraph(str(p.get('supplier_name','') or ''), small),
            Paragraph(str(p.get('supplier_gstin','') or '—'), small),
            str(p.get('bill_no','') or '—'),
            str(p.get('purchase_date','') or ''),
            Paragraph(str(p.get('place_of_supply','') or 'Same State'), small),
            rcm,
            _rupee(p.get('taxable_amount', p.get('total_amount', 0))),
            _rupee(p.get('cgst_amount', 0)),
            _rupee(p.get('sgst_amount', 0)),
            _rupee(p.get('igst_amount', 0)),
            _rupee(p.get('total_tax',   0)),
        ])

    # Totals row
    rows.append([
        '', 'TOTAL', '', '', '', '', '',
        _rupee(totals.get('taxable',   0)),
        _rupee(totals.get('cgst',      0)),
        _rupee(totals.get('sgst',      0)),
        _rupee(totals.get('igst',      0)),
        _rupee(totals.get('total_tax', 0)),
    ])

    col_w = [0.6*cm, 3.5*cm, 3.2*cm, 2.5*cm, 2.2*cm, 3*cm, 1.2*cm,
             2.8*cm, 2.3*cm, 2.3*cm, 2.3*cm, 2.5*cm]
    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    ts2 = _tbl_style()
    last = len(rows) - 1
    for col in range(7, 12):
        ts2.add('ALIGN', (col, 0), (col, -1), 'RIGHT')
    ts2.add('FONTNAME', (0, last), (-1, last), 'Helvetica-Bold')
    ts2.add('BACKGROUND', (0, last), (-1, last), rl_colors.HexColor('#EFF6FF'))
    tbl.setStyle(ts2)
    elems.append(tbl)

    elems.append(Spacer(1, 8))
    elems.append(Paragraph(
        '* RCM = Reverse Charge Mechanism. ITC from RCM purchases is separately claimable. '
        'Ensure all suppliers are GST-registered for ITC eligibility.',
        ParagraphStyle('note', fontName='Helvetica-Oblique', fontSize=7, textColor=C_MUTED)
    ))

    doc.build(elems)
    buf.seek(0)
    return buf.read()


def create_gstr2b_excel(data):
    """GSTR-2B style Purchase Register in Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Purchase Register'

    purchases = data.get('purchases', [])
    totals    = data.get('totals', {})
    co        = _load_settings().get('company_info', {})

    # ── Header ─────────────────────────────────────────────────
    ws.merge_cells('A1:L1')
    ws['A1'] = f"{co.get('name','Company')} — Purchase Register (ITC Statement)"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A2:L2')
    ws['A2'] = f"Period: {data['start_date']}  to  {data['end_date']}   |   GSTIN: {co.get('gstin','')}"
    ws['A2'].font = Font(size=9, italic=True)

    # ── Column Headers ──────────────────────────────────────────
    hdrs = ['#', 'Supplier Name', 'Supplier GSTIN', 'Bill No.', 'Bill Date',
            'Place of Supply', 'RCM', 'Taxable Value',
            'CGST', 'SGST', 'IGST', 'Total Tax']
    widths = [5, 22, 18, 14, 12, 18, 6, 15, 12, 12, 12, 14]

    def _w(row, col, val, bold=False, fill=None, align='left', fmt=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=9)
        if fill:
            c.fill = PatternFill('solid', fgColor=fill)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fmt:
            c.number_format = fmt

    ROW_HDR = 4
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _w(ROW_HDR, ci, h, bold=True, fill='1E40AF', align='center')
        ws.cell(row=ROW_HDR, column=ci).font = Font(bold=True, size=9, color='FFFFFF')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[ROW_HDR].height = 18
    ws.freeze_panes = f'A{ROW_HDR+1}'

    # ── Data rows ───────────────────────────────────────────────
    FILL_IGST = 'FFF7ED'
    FILL_ALT  = 'F8FAFC'
    r = ROW_HDR + 1
    for i, p in enumerate(purchases, 1):
        is_igst = float(p.get('igst_amount', 0) or 0) > 0
        bg = FILL_IGST if is_igst else (FILL_ALT if i % 2 == 0 else 'FFFFFF')
        _w(r, 1,  i,                                              align='center', fill=bg)
        _w(r, 2,  p.get('supplier_name',''),                      fill=bg, wrap=True)
        _w(r, 3,  p.get('supplier_gstin','') or '—',              fill=bg)
        _w(r, 4,  p.get('bill_no','') or '—',                     fill=bg)
        _w(r, 5,  p.get('purchase_date',''),                      fill=bg)
        _w(r, 6,  p.get('place_of_supply','') or 'Same State',    fill=bg)
        _w(r, 7,  'Yes' if p.get('reverse_charge') else 'No',    fill=bg, align='center')
        _w(r, 8,  float(p.get('taxable_amount', p.get('total_amount',0)) or 0), fill=bg, align='right', fmt='#,##0.00')
        _w(r, 9,  float(p.get('cgst_amount',0) or 0),             fill=bg, align='right', fmt='#,##0.00')
        _w(r, 10, float(p.get('sgst_amount',0) or 0),             fill=bg, align='right', fmt='#,##0.00')
        _w(r, 11, float(p.get('igst_amount',0) or 0),             fill=bg, align='right', fmt='#,##0.00')
        _w(r, 12, float(p.get('total_tax',  0) or 0),             fill=bg, align='right', fmt='#,##0.00')
        ws.row_dimensions[r].height = 16
        r += 1

    # ── Totals row ──────────────────────────────────────────────
    _w(r, 1, '',           bold=True, fill='DBEAFE')
    _w(r, 2, f'TOTAL ({totals.get("bill_count",0)} bills)', bold=True, fill='DBEAFE')
    for ci in range(3, 8):
        _w(r, ci, '', fill='DBEAFE')
    _w(r, 8,  totals.get('taxable',   0), bold=True, fill='DBEAFE', align='right', fmt='#,##0.00')
    _w(r, 9,  totals.get('cgst',      0), bold=True, fill='DBEAFE', align='right', fmt='#,##0.00')
    _w(r, 10, totals.get('sgst',      0), bold=True, fill='DBEAFE', align='right', fmt='#,##0.00')
    _w(r, 11, totals.get('igst',      0), bold=True, fill='DBEAFE', align='right', fmt='#,##0.00')
    _w(r, 12, totals.get('total_tax', 0), bold=True, fill='DBEAFE', align='right', fmt='#,##0.00')
    ws.row_dimensions[r].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
